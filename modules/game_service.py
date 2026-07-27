from datetime import datetime, date
import os
import openpyxl
import random
import sqlite3

# ดึงฟังก์ชันสุ่มไพ่ คำนวณคอมโบ และปรับ Luck จากโมดูลเดิม
from modules.card_engine import draw_cards
from modules.combo import check_combo
from modules.lucky import update_player_luck
from modules.score import calculate_score

# ---------------------------------------------------------
# Path Alignment (ชี้ไปยัง database/game.db และ Excel นอกโฟลเดอร์)
# ---------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB_PATH = os.path.join(BASE_DIR, "database", "game.db")
EXCEL_PATH = os.path.join(BASE_DIR, "CardGame-Sheet.xlsx")


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


# ---------------------------------------------------------
# Helper Functions
# ---------------------------------------------------------
def format_card_to_string(card):
    if isinstance(card, dict):
        rank = card.get("rank", card.get("value", ""))
        suit = card.get("suit", "")
        return f"{rank}{suit}".strip()
    return str(card).strip()


def evaluate_joker_combo(cards):
    """ฟังก์ชันตรวจเช็กและคำนวณคอมโบพิเศษสำหรับ Joker"""
    joker_cards = [c for c in cards if "Joker" in str(c)]
    joker_count = len(joker_cards)

    if joker_count == 0:
        return None

    if joker_count >= 2:
        return {
            "combo": "Double Joker 🃏🃏",
            "raw_score": 10,
            "play_cost": 0,
            "final_score": 10,
        }

    normal_cards = [c for c in cards if "Joker" not in str(c)]
    ranks = [
        c[:-1] if len(c) > 1 and c[-1] in ["♠", "♥", "♦", "♣"] else c
        for c in normal_cards
    ]

    if len(ranks) == 2 and ranks[0] == ranks[1]:
        return {
            "combo": "Wild Triple 🎰",
            "raw_score": 8,
            "play_cost": 0,
            "final_score": 8,
        }

    return {
        "combo": "Wild Pair 🃏✨",
        "raw_score": 5,
        "play_cost": 0,
        "final_score": 5,
    }


# ---------------------------------------------------------
# 1. ระบบตรวจสอบวันใหม่ และหักสิทธิ์เล่นฟรี
# ---------------------------------------------------------
def check_and_deduct_play(player_id):
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute(
        "SELECT daily_free, last_play_date, score, player_luck FROM players WHERE player_id = ?",
        (player_id,),
    )
    player = cursor.fetchone()

    if not player:
        conn.close()
        return False, "ไม่พบรหัสผู้เล่นในระบบ", None

    today_str = date.today().isoformat()
    last_date = (
        str(player["last_play_date"]).split(" ")[0]
        if player["last_play_date"]
        else ""
    )

    daily_free = player["daily_free"]

    # ข้ามวันใหม่ -> รีเซ็ตสิทธิ์ฟรีเป็น 2
    if last_date != today_str:
        daily_free = 2
        cursor.execute(
            "UPDATE players SET daily_free = 2, today_play = 0 WHERE player_id = ?",
            (player_id,),
        )

    if daily_free > 0:
        cursor.execute(
            """
            UPDATE players 
            SET daily_free = daily_free - 1, 
                today_play = today_play + 1,
                last_play_date = ? 
            WHERE player_id = ?
        """,
            (today_str, player_id),
        )
        conn.commit()
        conn.close()
        return True, "หักสิทธิ์สำเร็จ", player
    else:
        conn.close()
        return False, "สิทธิ์เล่นฟรีประจำวันหมดแล้ว", player


# ---------------------------------------------------------
# 2. ระบบ Sync คะแนนและเวลาลงไฟล์ Excel
# ---------------------------------------------------------
def sync_to_excel(player_id, score_gained):
    if not os.path.exists(EXCEL_PATH):
        return

    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = [cell.value for cell in ws[1]]

            if "รหัสตัวละคร" in headers and "Total Score" in headers:
                id_col = headers.index("รหัสตัวละคร") + 1
                score_col = headers.index("Total Score") + 1
                date_col = (
                    headers.index("Last Play Date") + 1
                    if "Last Play Date" in headers
                    else None
                )
                free_col = (
                    headers.index("Free Plays") + 1
                    if "Free Plays" in headers
                    else None
                )

                for row in range(2, ws.max_row + 1):
                    cell_val = ws.cell(row=row, column=id_col).value
                    if (
                        cell_val
                        and str(cell_val).strip() == str(player_id).strip()
                    ):
                        current_score = (
                            ws.cell(row=row, column=score_col).value or 0
                        )
                        ws.cell(
                            row=row,
                            column=score_col,
                            value=int(current_score) + score_gained,
                        )

                        if date_col:
                            ws.cell(row=row, column=date_col, value=now_str)

                        if free_col:
                            current_free = (
                                ws.cell(row=row, column=free_col).value or 2
                            )
                            ws.cell(
                                row=row,
                                column=free_col,
                                value=max(0, int(current_free) - 1),
                            )

                        wb.save(EXCEL_PATH)
                        return
    except Exception as e:
        print(f"⚠️ Sync Excel Error: {e}")


# ---------------------------------------------------------
# 3. ฟังก์ชันหลัก: เล่นเกม + หักสิทธิ์ + ลง DB + Sync Excel (เรียกอันนี้อันเดียว)
# ---------------------------------------------------------
def play_game(player_id, event_luck=0.0):
    # Step 1: ตรวจสิทธิ์และหักสิทธิ์
    can_play, msg, player_info = check_and_deduct_play(player_id)
    if not can_play:
        return {"success": False, "message": msg}

    player_score = player_info["score"] if player_info else 0
    player_luck = player_info["player_luck"] if player_info else 0.0

    # Step 2: คำนวณ Luck และสุ่มไพ่ 3 ใบ
    final_luck = round(player_luck + event_luck, 2)
    try:
        raw_cards = draw_cards(luck=final_luck)
        cards = (
            [format_card_to_string(c) for c in raw_cards]
            if isinstance(raw_cards, list)
            else [format_card_to_string(raw_cards)]
        )
    except Exception as e:
        print(f"⚠️ Card Engine Error: {e}")
        suits = ["♠", "♥", "♦", "♣"]
        ranks = [
            "2",
            "3",
            "4",
            "5",
            "6",
            "7",
            "8",
            "9",
            "10",
            "J",
            "Q",
            "K",
            "A",
        ]
        cards = [
            f"{random.choice(ranks)}{random.choice(suits)}" for _ in range(3)
        ]

    # Step 3: คำนวณคอมโบและคะแนน
    joker_result = evaluate_joker_combo(cards)
    if joker_result:
        combo_name = joker_result["combo"]
        score_gained = joker_result["raw_score"]
        play_cost = joker_result["play_cost"]
    else:
        try:
            combo_name = check_combo(cards)
        except Exception:
            combo_name = "High Card"

        try:
            raw_score, play_cost, final_calc_score, _ = calculate_score(
                combo_name, player_score
            )
            score_gained = raw_score
        except Exception:
            score_gained = 0
            play_cost = 0

    # Step 4: ปรับค่า Player Luck
    try:
        next_luck = update_player_luck(
            current_luck=player_luck, combo=combo_name, cards=cards
        )
    except Exception:
        next_luck = player_luck

    # Step 5: บันทึกลง SQLite (game_history และ อัปเดต score/luck ใน players)
    total_score = player_score + score_gained
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute(
        "UPDATE players SET score = ?, player_luck = ? WHERE player_id = ?",
        (total_score, next_luck, player_id),
    )

    cards_str = ", ".join(cards)
    cursor.execute(
        """
        INSERT INTO game_history (player_id, cards_drawn, combo_name, score_gained, final_score)
        VALUES (?, ?, ?, ?, ?)
    """,
        (player_id, cards_str, combo_name, score_gained, total_score),
    )

    conn.commit()
    conn.close()

    # Step 6: Sync ลง Excel
    sync_to_excel(player_id, score_gained)

    # ส่งผลลัพธ์กลับไปใช้งาน
    return {
        "success": True,
        "player_id": player_id,
        "cards": cards,
        "combo": combo_name,
        "score_gained": score_gained,
        "cost": play_cost,
        "final_score": total_score,
        "player_luck": player_luck,
        "next_player_luck": next_luck,
    }
